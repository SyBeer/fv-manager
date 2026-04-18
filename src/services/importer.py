"""Import monthly readings from Excel (Moja instalacja sheet)."""
import openpyxl
from pathlib import Path
from dataclasses import dataclass, field

# Column indices in "Moja instalacja" sheet (0-based, row 5+ = data)
COL_PERIOD        = 0
COL_DAYS          = 1
COL_PRODUCTION    = 3   # Produkcja SolarEdge
COL_SENT_METER    = 4   # Oddane 2.8.0 - Licznik (cumulative)
COL_TAKEN_METER   = 5   # Pobrane 1.8.0 - Licznik (cumulative)
COL_SENT_CALC     = 6   # Oddane obliczenia (monthly delta)
COL_TAKEN_CALC    = 8   # Pobrane obliczenia (monthly delta)
COL_INVOICE_NUM   = 16  # Nr Faktury PGE
COL_PRICE_PER_KWH = 21  # Cena za kWh z faktury
COL_INVOICE_GROSS = 26  # FAKTURA brutto


@dataclass
class ImportResult:
    records: list[dict] = field(default_factory=list)
    rejected: list[dict] = field(default_factory=list)  # sent > production

    @property
    def has_rejections(self) -> bool:
        return len(self.rejected) > 0


def import_excel(path: str | Path) -> ImportResult:
    wb = openpyxl.load_workbook(path, data_only=True)

    if "Moja instalacja" not in wb.sheetnames:
        raise ValueError(
            f"Brak arkusza 'Moja instalacja'. "
            f"Dostępne arkusze: {wb.sheetnames}"
        )

    ws = wb["Moja instalacja"]
    result = ImportResult()

    for row in ws.iter_rows(min_row=5, max_row=200, values_only=True):
        period = row[COL_PERIOD]
        if not period or not isinstance(period, str):
            continue

        production = row[COL_PRODUCTION]
        sent_meter = row[COL_SENT_METER]
        taken_meter = row[COL_TAKEN_METER]
        sent_calc = row[COL_SENT_CALC]
        taken_calc = row[COL_TAKEN_CALC]

        if production is None:
            continue

        # Only import rows with real meter data (not future projections)
        if sent_meter is None or taken_meter is None:
            continue

        sent = float(sent_calc) if sent_calc is not None else 0.0
        taken = float(taken_calc) if taken_calc is not None else 0.0
        prod = float(production)

        try:
            year, month = int(period.split(".")[0]), int(period.split(".")[1])
        except (ValueError, IndexError):
            continue

        # Reject physically impossible records
        if sent > prod:
            result.rejected.append({
                "period": period,
                "production_kwh": round(prod, 3),
                "sent_to_grid_kwh": round(sent, 3),
                "reason": f"Oddane ({sent:.1f} kWh) > Produkcja ({prod:.1f} kWh)",
            })
            continue

        result.records.append({
            "period": period,
            "year": year,
            "month": month,
            "days": row[COL_DAYS],
            "production_kwh": round(prod, 3),
            "sent_to_grid_kwh": round(sent, 3),
            "taken_from_grid_kwh": round(taken, 3),
            "price_per_kwh": row[COL_PRICE_PER_KWH],
            "invoice_number": row[COL_INVOICE_NUM],
            "invoice_gross": row[COL_INVOICE_GROSS],
        })

    return result
